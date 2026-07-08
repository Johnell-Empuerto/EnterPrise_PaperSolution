"""
Phase 1.2 - Test Workbook Generator
Creates a comprehensive set of Excel workbooks to validate the Print Area capture pipeline.
All files are saved to the test_files_output/ directory.
"""

import os
import io
import struct
import zlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_files_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def save_with_print_area(wb, filename, print_area="A1:J30"):
    """Set print area on first worksheet and save."""
    ws = wb.active
    # NOTE: print_area is a property on Worksheet, not on page_setup
    ws.print_area = print_area
    ws.page_setup.orientation = "portrait"
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  Created: {filename}")
    return path


# ============================================================
# 1. Simple Table
# ============================================================
def test_01_simple_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Simple Table"
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=2, value="Name")
    ws.cell(row=1, column=3, value="Amount")
    ws.cell(row=1, column=4, value="Date")
    data = [
        (1, "Alice Johnson", 1250.50, "2025-01-15"),
        (2, "Bob Smith", 780.00, "2025-02-20"),
        (3, "Carol Davis", 2340.75, "2025-03-10"),
        (4, "David Wilson", 560.25, "2025-04-05"),
        (5, "Eve Brown", 1890.00, "2025-05-18"),
        (6, "Frank Miller", 450.00, "2025-06-22"),
        (7, "Grace Lee", 3100.00, "2025-07-14"),
        (8, "Henry Taylor", 920.50, "2025-08-30"),
    ]
    for i, (id_, name, amt, date) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=id_)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=amt)
        ws.cell(row=i, column=4, value=date)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    return save_with_print_area(wb, "01_simple_table.xlsx", "A1:D12")


# ============================================================
# 2. Borders (Thick, Thin, Mixed)
# ============================================================
def test_02_borders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Borders"
    thin_side = Side(style="thin", color="000000")
    thick_side = Side(style="thick", color="000000")
    double_side = Side(style="double", color="000000")
    dashed_side = Side(style="dashed", color="333333")
    dotted_side = Side(style="dotted", color="666666")

    styles = [
        ("Thin All", Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)),
        ("Thick All", Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)),
        ("Thin Left/Right", Border(left=thin_side, right=thin_side)),
        ("Double Bottom", Border(bottom=double_side)),
        ("Dashed Outline", Border(left=dashed_side, right=dashed_side, top=thin_side, bottom=thin_side)),
        ("Dotted All", Border(left=dotted_side, right=dotted_side, top=dotted_side, bottom=dotted_side)),
        ("Mixed", Border(left=thick_side, right=thin_side, top=double_side, bottom=dashed_side)),
        ("No Border", Border()),
    ]

    ws.cell(row=1, column=1, value="Style Name")
    ws.cell(row=1, column=2, value="Sample Cell")
    for cell in [ws.cell(row=1, column=1), ws.cell(row=1, column=2)]:
        cell.font = Font(bold=True, size=12)

    for i, (name, border) in enumerate(styles, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value="Hello").border = border
        ws.cell(row=i, column=1).border = border

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.page_setup.print_area = "A1:B10"
    return save_with_print_area(wb, "02_borders.xlsx", "A1:B10")


# ============================================================
# 3. Merged Cells
# ============================================================
def test_03_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged Cells"

    ws.cell(row=1, column=1, value="MAIN HEADER SPANNING MULTIPLE COLUMNS").font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.cell(row=3, column=1, value="Left Info")
    ws.merge_cells("A3:B3")
    ws.cell(row=3, column=3, value="Right Info")
    ws.merge_cells("C3:D3")

    ws.cell(row=5, column=1, value="Vertical Merge")
    ws.merge_cells("A5:A8")
    ws.cell(row=5, column=1).alignment = Alignment(vertical="center", horizontal="center")
    ws.cell(row=5, column=1).font = Font(bold=True, size=14)

    for r in range(5, 9):
        ws.cell(row=r, column=2, value=f"Row {r} data")

    ws.cell(row=5, column=3, value="Multi-cell merge area")
    ws.merge_cells("C5:D8")
    ws.cell(row=5, column=3).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 22
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[5].height = 60

    return save_with_print_area(wb, "03_merged_cells.xlsx", "A1:D10")


# ============================================================
# 4. Wrapped Text
# ============================================================
def test_04_wrapped_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Wrapped Text"

    headers = ["Short", "Very Long Wrapped Text", "Mixed Length Content", "Notes"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    long_texts = [
        "This is a very long sentence that should wrap across multiple lines inside the cell because it exceeds the column width.",
        "Short text.",
        "Medium length wrapping here too.",
        "Another lengthy paragraph that will demonstrate how Excel wraps text when the column width is insufficient to show the entire content on one line.",
    ]

    for r in range(2, 12):
        for c, t in enumerate(long_texts, start=1):
            cell = ws.cell(row=r, column=c, value=f"{t} Row {r}")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    for r in range(2, 12):
        ws.row_dimensions[r].height = 45

    return save_with_print_area(wb, "04_wrapped_text.xlsx", "A1:D12")


# ============================================================
# 5. Different Fonts
# ============================================================
def test_05_fonts():
    wb = Workbook()
    ws = wb.active
    ws.title = "Fonts"

    fonts = [
        ("Arial", "Calibri"),
        ("Times New Roman", "Verdana"),
        ("Courier New", "Georgia"),
        ("Tahoma", "Trebuchet MS"),
        ("Consolas", "Segoe UI"),
    ]

    ws.cell(row=1, column=1, value="Font Family").font = Font(bold=True, size=11)
    ws.cell(row=1, column=2, value="Sample Text").font = Font(bold=True, size=11)

    for i, (font1, font2) in enumerate(fonts, start=2):
        ws.cell(row=i, column=1, value=font1).font = Font(name=font1, size=14)
        ws.cell(row=i, column=2, value=f"Hello World in {font1}").font = Font(name=font1, size=14)
        ws.cell(row=i + 6, column=1, value=font2).font = Font(name=font2, size=14)
        ws.cell(row=i + 6, column=2, value=f"Hello World in {font2}").font = Font(name=font2, size=14)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 35
    return save_with_print_area(wb, "05_fonts.xlsx", "A1:B16")


# ============================================================
# 6. Font Sizes and Colors
# ============================================================
def test_06_font_sizes_colors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Font Sizes & Colors"

    sizes = [8, 10, 12, 14, 16, 18, 20, 24, 28, 36]
    ws.cell(row=1, column=1, value="Size").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Sample").font = Font(bold=True)
    ws.cell(row=1, column=3, value="Color").font = Font(bold=True)
    ws.cell(row=1, column=4, value="Colored Sample").font = Font(bold=True)

    colors = [
        ("FF0000", "Red"),
        ("0000FF", "Blue"),
        ("008000", "Green"),
        ("800080", "Purple"),
        ("FF8C00", "Orange"),
        ("000000", "Black"),
    ]

    for i, size in enumerate(sizes, start=2):
        ws.cell(row=i, column=1, value=size)
        ws.cell(row=i, column=2, value=f"Font size {size}").font = Font(size=size)

    for i, (color, name) in enumerate(colors, start=2):
        ws.cell(row=i, column=3, value=name).font = Font(size=12)
        ws.cell(row=i, column=4, value=f"Colored text").font = Font(size=14, color=color)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    return save_with_print_area(wb, "06_font_sizes_colors.xlsx", "A1:D14")


# ============================================================
# 7. Background Fills
# ============================================================
def test_07_background_fills():
    wb = Workbook()
    ws = wb.active
    ws.title = "Background Fills"

    fills = [
        ("Yellow", PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")),
        ("Light Blue", PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")),
        ("Light Green", PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")),
        ("Light Gray", PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")),
        ("Orange", PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")),
        ("Lavender", PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")),
        ("Pink", PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")),
        ("Dark Header", PatternFill(start_color="333333", end_color="333333", fill_type="solid")),
    ]

    ws.cell(row=1, column=1, value="Fill Name").font = Font(bold=True, size=12)
    ws.cell(row=1, column=2, value="Sample").font = Font(bold=True, size=12)
    ws.cell(row=1, column=3, value="With Text").font = Font(bold=True, size=12)

    for i, (name, fill) in enumerate(fills, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value="").fill = fill
        ws.cell(row=i, column=3, value=f"Filled {name}").fill = fill
        text_color = "FFFFFF" if name == "Dark Header" else "000000"
        ws.cell(row=i, column=3).font = Font(color=text_color, bold=True)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    return save_with_print_area(wb, "07_background_fills.xlsx", "A1:C12")


# ============================================================
# 8. Shapes (via comments and rich text) — no native shape support in openpyxl
# ============================================================
def test_08_shapes():
    """Create a worksheet with shapes (approximated via merged cells and borders)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shapes"

    # Simulate a "button" shape using merged cells, borders, and fill
    ws.cell(row=2, column=2, value="CLICK ME").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=2, column=2).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B2:C3")
    thin_side = Side(style="thin", color="2F5496")
    ws.cell(row=2, column=2).border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25

    # Simulate a "text box" area
    ws.cell(row=5, column=2, value="This simulates a text box area with a border and fill.").font = Font(size=11)
    ws.cell(row=5, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("B5:E7")
    ws.cell(row=5, column=2).border = Border(
        left=Side(style="medium", color="999999"),
        right=Side(style="medium", color="999999"),
        top=Side(style="medium", color="999999"),
        bottom=Side(style="medium", color="999999"),
    )
    ws.cell(row=5, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 20

    # Simulate a "line" via a thin filled row
    ws.cell(row=9, column=2, value="").fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws.merge_cells("B9:E9")
    ws.row_dimensions[9].height = 3

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    return save_with_print_area(wb, "08_shapes.xlsx", "A1:E12")


# ============================================================
# 9. Hidden Rows
# ============================================================
def test_09_hidden_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "Hidden Rows"

    headers = ["Visible?", "Item", "Value", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True, size=12)

    for r in range(2, 12):
        ws.cell(row=r, column=1, value="Yes" if r not in (4, 7, 9) else "No")
        ws.cell(row=r, column=2, value=f"Item {r-1}")
        ws.cell(row=r, column=3, value=r * 100)
        ws.cell(row=r, column=4, value=f"Row {r} notes")
        if r in (4, 7, 9):
            ws.row_dimensions[r].hidden = True

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    return save_with_print_area(wb, "09_hidden_rows.xlsx", "A1:D12")


# ============================================================
# 10. Hidden Columns
# ============================================================
def test_10_hidden_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "Hidden Columns"

    headers = ["Visible", "Hidden", "Visible", "Hidden2", "Visible"]
    labels = ["Col A", "Col B", "Col C", "Col D", "Col E"]
    for c, (h, l) in enumerate(zip(headers, labels), start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True, size=11)
        ws.cell(row=2, column=c, value=l)
        ws.cell(row=3, column=c, value=f"Data {c}")

    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["D"].hidden = True
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    return save_with_print_area(wb, "10_hidden_columns.xlsx", "A1:E6")


# ============================================================
# 11. Frozen Panes
# ============================================================
def test_11_frozen_panes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Frozen Panes"
    ws.freeze_panes = "B3"

    for c in range(1, 8):
        ws.cell(row=1, column=c, value=f"Header {get_column_letter(c)}").font = Font(bold=True, size=11)
    for r in range(2, 20):
        ws.cell(row=r, column=1, value=f"Row {r}").font = Font(bold=True, size=11)
        for c in range(2, 8):
            ws.cell(row=r, column=c, value=r * c)

    for c in range(1, 8):
        ws.column_dimensions[get_column_letter(c)].width = 12

    return save_with_print_area(wb, "11_frozen_panes.xlsx", "A1:G12")


# ============================================================
# 12. Variable Row Heights and Column Widths
# ============================================================
def test_12_variable_sizes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Variable Sizes"

    heights = [15, 25, 35, 50, 70, 20, 40, 60, 30, 45]
    widths = [8, 15, 25, 35, 12, 20, 30, 18, 10, 40]

    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
        ws.cell(row=1, column=c, value=f"Col {c}").font = Font(bold=True, size=10)

    for r, h in enumerate(heights, start=2):
        ws.row_dimensions[r].height = h
        for c in range(1, 8):
            ws.cell(row=r, column=c, value=f"R{r}C{c}")

    return save_with_print_area(wb, "12_variable_sizes.xlsx", "A1:G12")


# ============================================================
# 13. Landscape Orientation
# ============================================================
def test_13_landscape():
    wb = Workbook()
    ws = wb.active
    ws.title = "Landscape"
    ws.page_setup.orientation = "landscape"

    for c in range(1, 12):
        ws.cell(row=1, column=c, value=f"Header {c}").font = Font(bold=True, size=10)
    for r in range(2, 20):
        for c in range(1, 12):
            ws.cell(row=r, column=c, value=f"R{r}C{c}")

    for c in range(1, 12):
        ws.column_dimensions[get_column_letter(c)].width = 14

    return save_with_print_area(wb, "13_landscape.xlsx", "A1:K15")


# ============================================================
# 14. Portrait Orientation with Fit to Page
# ============================================================
def test_14_fit_to_page():
    wb = Workbook()
    ws = wb.active
    ws.title = "Fit to Page"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    for c in range(1, 10):
        ws.cell(row=1, column=c, value=f"Header {c}").font = Font(bold=True, size=10)
    for r in range(2, 35):
        for c in range(1, 10):
            ws.cell(row=r, column=c, value=f"Data {r},{c}")

    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 12

    return save_with_print_area(wb, "14_fit_to_page.xlsx", "A1:I30")


# ============================================================
# 15. Multiple Print Areas (openpyxl sets only one print area)
# ============================================================
def test_15_multiple_areas():
    """Test with ranges including commas to simulate multiple areas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi Area"

    # Write data in two separate areas
    for r in range(1, 6):
        for c in range(1, 4):
            ws.cell(row=r, column=c, value=f"Area1-R{r}C{c}")
    for r in range(1, 6):
        for c in range(6, 9):
            ws.cell(row=r, column=c, value=f"Area2-R{r}C{c}")

    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # openpyxl stores print_area as a string; we can set it to multiple ranges
    return save_with_print_area(wb, "15_multiple_areas.xlsx", "A1:C5,F1:H5")


# ============================================================
# 16. Large Print Area
# ============================================================
def test_16_large_area():
    wb = Workbook()
    ws = wb.active
    ws.title = "Large Area"

    for r in range(1, 51):
        for c in range(1, 21):
            ws.cell(row=r, column=c, value=f"D{r}C{c}")

    for c in range(1, 21):
        ws.column_dimensions[get_column_letter(c)].width = 10

    return save_with_print_area(wb, "16_large_area.xlsx", "A1:T50")


# ============================================================
# 17. Multiple Worksheets
# ============================================================
def test_17_multiple_worksheets():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.cell(row=1, column=1, value="Sheet 1 Data")
    ws1.cell(row=2, column=1, value="This should be captured")

    ws2 = wb.create_sheet("Sheet2")
    ws2.cell(row=1, column=1, value="Sheet 2 Data")
    ws2.cell(row=2, column=1, value="This should NOT appear")

    ws3 = wb.create_sheet("Sheet3")
    ws3.cell(row=1, column=1, value="Sheet 3 Data")
    ws3.cell(row=2, column=1, value="This should NOT appear either")

    ws1.column_dimensions["A"].width = 35
    ws2.column_dimensions["A"].width = 35
    ws3.column_dimensions["A"].width = 35

    return save_with_print_area(wb, "17_multiple_worksheets.xlsx", "A1:B5")


# ============================================================
# 18. Print Titles (Rows to repeat at top)
# ============================================================
def test_18_print_titles():
    wb = Workbook()
    ws = wb.active
    ws.title = "Print Titles"
    ws.print_title_rows = "1:2"

    headers = ["ID", "Product", "Description", "Quantity", "Price", "Total"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True, size=11)
        ws.cell(row=2, column=c, value=f"({h} subtitle)").font = Font(size=9, italic=True)

    for r in range(3, 12):
        ws.cell(row=r, column=1, value=r - 2)
        ws.cell(row=r, column=2, value=f"Product {r - 2}")
        ws.cell(row=r, column=3, value=f"Description for item {r - 2}")
        ws.cell(row=r, column=4, value=r * 5)
        ws.cell(row=r, column=5, value=round(r * 12.5, 2))
        ws.cell(row=r, column=6, value=round(r * 5 * r * 12.5, 2))

    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 16

    return save_with_print_area(wb, "18_print_titles.xlsx", "A1:F12")


# ============================================================
# 19. Headers and Footers
# ============================================================
def test_19_headers_footers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Headers & Footers"
    # Note: openpyxl may not support header_footer in all versions
    # Try setting it, but accept if it fails
    try:
        ws.oddHeader.left.text = "PaperLess Enterprise"
        ws.oddHeader.center.text = "&[File]"
        ws.oddHeader.right.text = "Page &[Page] of &[Pages]"
        ws.oddFooter.left.text = "Confidential"
        ws.oddFooter.center.text = "Prepared on &[Date]"
        ws.oddFooter.right.text = "&[Tab]"
    except AttributeError:
        pass

    ws.cell(row=1, column=1, value="Header/Footer Test").font = Font(bold=True, size=16)
    ws.cell(row=3, column=1, value="Check that the image shows the print area content.")
    ws.cell(row=4, column=1, value="Headers/footers typically do NOT appear in the captured PNG.")

    ws.column_dimensions["A"].width = 55

    return save_with_print_area(wb, "19_headers_footers.xlsx", "A1:B6")


# ============================================================
# 20. Empty Print Area (Should trigger error)
# ============================================================
def test_20_empty_print_area():
    """Create a workbook with NO print area set — should trigger the 400 error."""
    wb = Workbook()
    ws = wb.active
    ws.title = "No Print Area"
    ws.cell(row=1, column=1, value="This worksheet has no print area configured.")
    ws.cell(row=2, column=1, value="Uploading this should return a 400 error.")
    ws.column_dimensions["A"].width = 55
    path = os.path.join(OUTPUT_DIR, "20_empty_print_area.xlsx")
    wb.save(path)
    print(f"  Created: 20_empty_print_area.xlsx")
    return path


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    print("Generating Phase 1.2 test workbooks...")
    print(f"Output directory: {OUTPUT_DIR}\n")

    test_files = [
        ("01_simple_table", test_01_simple_table),
        ("02_borders", test_02_borders),
        ("03_merged_cells", test_03_merged_cells),
        ("04_wrapped_text", test_04_wrapped_text),
        ("05_fonts", test_05_fonts),
        ("06_font_sizes_colors", test_06_font_sizes_colors),
        ("07_background_fills", test_07_background_fills),
        ("08_shapes", test_08_shapes),
        ("09_hidden_rows", test_09_hidden_rows),
        ("10_hidden_columns", test_10_hidden_columns),
        ("11_frozen_panes", test_11_frozen_panes),
        ("12_variable_sizes", test_12_variable_sizes),
        ("13_landscape", test_13_landscape),
        ("14_fit_to_page", test_14_fit_to_page),
        ("15_multiple_areas", test_15_multiple_areas),
        ("16_large_area", test_16_large_area),
        ("17_multiple_worksheets", test_17_multiple_worksheets),
        ("18_print_titles", test_18_print_titles),
        ("19_headers_footers", test_19_headers_footers),
        ("20_empty_print_area", test_20_empty_print_area),
    ]

    for name, func in test_files:
        try:
            func()
        except Exception as e:
            print(f"  ERROR creating {name}: {e}")

    print(f"\nAll test files generated in: {OUTPUT_DIR}")
    print(f"Total files: {len([f for f in os.listdir(OUTPUT_DIR) if f.endswith('.xlsx')])}")
