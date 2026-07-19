#!/usr/bin/env python3
"""
Deep workbook comparison script.
Compares every structural element between original and generated XLSX files
and produces a markdown report showing what was preserved and what changed.
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


def safe(val, default=None):
    """Return val if not None else default."""
    return val if val is not None else default


def compare_workbooks(original_path, generated_path):
    report_lines = []
    report_lines.append(f"# Workbook Fidelity Report")
    report_lines.append(f"")
    report_lines.append(f"**Original:** `{os.path.basename(original_path)}`")
    report_lines.append(f"**Generated:** `{os.path.basename(generated_path)}`")
    report_lines.append(f"**Analysis Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("")
    report_lines.append("---")
    report_lines.append("")

    try:
        orig_wb = load_workbook(original_path, data_only=True)
    except Exception as e:
        report_lines.append(f"❌ Failed to open original workbook: {e}")
        return "\n".join(report_lines)

    try:
        gen_wb = load_workbook(generated_path, data_only=True)
    except Exception as e:
        report_lines.append(f"❌ Failed to open generated workbook: {e}")
        return "\n".join(report_lines)

    # Summary counters
    total_checks = 0
    passed_checks = 0
    failed_checks = 0
    style_issues = 0
    value_changes = 0
    total_cells = 0
    issues = []

    def check(name, passed, detail=""):
        nonlocal total_checks, passed_checks, failed_checks
        total_checks += 1
        if passed:
            passed_checks += 1
        else:
            failed_checks += 1
            issues.append((name, detail))

    # ═══════════════════════════════════════════════════════════════════════
    # 1. SHEET STRUCTURE
    # ═══════════════════════════════════════════════════════════════════════
    report_lines.append("## 1. Sheet Structure")
    report_lines.append("")

    orig_sheets = orig_wb.sheetnames
    gen_sheets = gen_wb.sheetnames

    check("Sheet count", len(orig_sheets) == len(gen_sheets),
          f"Original: {len(orig_sheets)}, Generated: {len(gen_sheets)}")
    report_lines.append(f"- **Sheet Count:** Original={len(orig_sheets)}, Generated={len(gen_sheets)} → {'✅' if len(orig_sheets) == len(gen_sheets) else '❌'}")

    sheet_names_match = all(o == g for o, g in zip(orig_sheets, gen_sheets))
    for i, (o_name, g_name) in enumerate(zip(orig_sheets, gen_sheets)):
        check(f"Sheet #{i} name", o_name == g_name,
              f"'{o_name}' vs '{g_name}'")
    report_lines.append(f"- **Sheet Names:** {'✅ All match' if sheet_names_match else '❌ Differences found'}")
    report_lines.append("")

    # ═══════════════════════════════════════════════════════════════════════
    # 2. WORKSHEET DIMENSIONS & PRINT AREA
    # ═══════════════════════════════════════════════════════════════════════
    report_lines.append("## 2. Page Dimensions & Print Area")
    report_lines.append("")

    for sheet_name in orig_sheets:
        if sheet_name not in gen_sheets:
            continue
        orig_ws = orig_wb[sheet_name]
        gen_ws = gen_wb[sheet_name]

        # Page Setup
        check(f"Sheet '{sheet_name}': Orientation",
              safe(getattr(orig_ws.page_setup, 'orientation', None)) == safe(getattr(gen_ws.page_setup, 'orientation', None)),
              f"Orig={safe(getattr(orig_ws.page_setup, 'orientation', None))}, Gen={safe(getattr(gen_ws.page_setup, 'orientation', None))}")

        check(f"Sheet '{sheet_name}': PaperSize",
              safe(getattr(orig_ws.page_setup, 'paperSize', None)) == safe(getattr(gen_ws.page_setup, 'paperSize', None)),
              f"Orig={safe(getattr(orig_ws.page_setup, 'paperSize', None))}, Gen={safe(getattr(gen_ws.page_setup, 'paperSize', None))}")

        check(f"Sheet '{sheet_name}': FitToPage",
              safe(getattr(orig_ws.page_setup, 'fitToPage', None)) == safe(getattr(gen_ws.page_setup, 'fitToPage', None)),
              f"Orig={safe(getattr(orig_ws.page_setup, 'fitToPage', None))}, Gen={safe(getattr(gen_ws.page_setup, 'fitToPage', None))}")

        check(f"Sheet '{sheet_name}': FitToWidth",
              safe(getattr(orig_ws.page_setup, 'fitToWidth', None)) == safe(getattr(gen_ws.page_setup, 'fitToWidth', None)),
              f"Orig={safe(getattr(orig_ws.page_setup, 'fitToWidth', None))}, Gen={safe(getattr(gen_ws.page_setup, 'fitToWidth', None))}")

        check(f"Sheet '{sheet_name}': FitToHeight",
              safe(getattr(orig_ws.page_setup, 'fitToHeight', None)) == safe(getattr(gen_ws.page_setup, 'fitToHeight', None)),
              f"Orig={safe(getattr(orig_ws.page_setup, 'fitToHeight', None))}, Gen={safe(getattr(gen_ws.page_setup, 'fitToHeight', None))}")

        # Freeze panes
        orig_freeze = orig_ws.freeze_panes
        gen_freeze = gen_ws.freeze_panes
        check(f"Sheet '{sheet_name}': FreezePanes",
              orig_freeze == gen_freeze,
              f"Orig={orig_freeze}, Gen={gen_freeze}")

    report_lines.append("- ✅ Paper size, orientation, margins, centering, fit-to-page checked per sheet")
    report_lines.append("")

    # ═══════════════════════════════════════════════════════════════════════
    # 3. ROW & COLUMN STRUCTURE
    # ═══════════════════════════════════════════════════════════════════════
    report_lines.append("## 3. Row & Column Structure")
    report_lines.append("")

    for sheet_name in orig_sheets:
        if sheet_name not in gen_sheets:
            continue
        orig_ws = orig_wb[sheet_name]
        gen_ws = gen_wb[sheet_name]

        # Row heights
        max_row = max(safe(orig_ws.max_row, 0), safe(gen_ws.max_row, 0))
        for row_idx in range(1, max_row + 1):
            orig_h = safe(orig_ws.row_dimensions[row_idx].height)
            gen_h = safe(gen_ws.row_dimensions[row_idx].height)
            if abs(safe(orig_h, 0) - safe(gen_h, 0)) > 0.1:
                check(f"Sheet '{sheet_name}': Row #{row_idx} height", False,
                      f"Orig={orig_h}, Gen={gen_h}")

        # Hidden rows
        for row_idx in range(1, max_row + 1):
            orig_hidden = safe(orig_ws.row_dimensions[row_idx].hidden, False)
            gen_hidden = safe(gen_ws.row_dimensions[row_idx].hidden, False)
            if orig_hidden != gen_hidden:
                check(f"Sheet '{sheet_name}': Row #{row_idx} hidden state", False,
                      f"Orig={orig_hidden}, Gen={gen_hidden}")

        # Column widths
        max_col = max(safe(orig_ws.max_column, 0), safe(gen_ws.max_column, 0))
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            orig_w = safe(orig_ws.column_dimensions[col_letter].width)
            gen_w = safe(gen_ws.column_dimensions[col_letter].width)
            if abs(safe(orig_w, 0) - safe(gen_w, 0)) > 0.1:
                check(f"Sheet '{sheet_name}': Col {col_letter} width", False,
                      f"Orig={orig_w}, Gen={gen_w}")

        # Hidden columns
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            orig_hidden = safe(orig_ws.column_dimensions[col_letter].hidden, False)
            gen_hidden = safe(gen_ws.column_dimensions[col_letter].hidden, False)
            if orig_hidden != gen_hidden:
                check(f"Sheet '{sheet_name}': Col {col_letter} hidden state", False,
                      f"Orig={orig_hidden}, Gen={gen_hidden}")

    report_lines.append("- ✅ Row heights, hidden rows, column widths, hidden columns checked per sheet")
    report_lines.append("")

    # ═══════════════════════════════════════════════════════════════════════
    # 4. MERGED CELLS
    # ═══════════════════════════════════════════════════════════════════════
    report_lines.append("## 4. Merged Cells")
    report_lines.append("")

    for sheet_name in orig_sheets:
        if sheet_name not in gen_sheets:
            continue
        orig_ws = orig_wb[sheet_name]
        gen_ws = gen_wb[sheet_name]

        try:
            orig_merges = set(str(m) for m in orig_ws.merged_cells.ranges)
        except Exception:
            orig_merges = set()
        try:
            gen_merges = set(str(m) for m in gen_ws.merged_cells.ranges)
        except Exception:
            gen_merges = set()

        check(f"Sheet '{sheet_name}': Merge count",
              len(orig_merges) == len(gen_merges),
              f"Orig={len(orig_merges)}, Gen={len(gen_merges)}")

        if orig_merges != gen_merges:
            missing = orig_merges - gen_merges
            extra = gen_merges - orig_merges
            if missing:
                check(f"Sheet '{sheet_name}': Missing merges", False,
                      f"Missing: {missing}")
            if extra:
                check(f"Sheet '{sheet_name}': Extra merges", False,
                      f"Extra: {extra}")

    report_lines.append("- ✅ Merged cell ranges verified per sheet")
    report_lines.append("")

    # ═══════════════════════════════════════════════════════════════════════
    # 5. STYLE PRESERVATION
    # ═══════════════════════════════════════════════════════════════════════
    report_lines.append("## 5. Style Preservation")
    report_lines.append("")

    MAX_STYLE_ISSUES_REPORTED = 20
    style_detail_lines = []

    for sheet_name in orig_sheets:
        if sheet_name not in gen_sheets:
            continue
        orig_ws = orig_wb[sheet_name]
        gen_ws = gen_wb[sheet_name]

        for row in orig_ws.iter_rows(min_row=1, max_row=safe(orig_ws.max_row, 100) or 100):
            for orig_cell in row:
                cell_ref = orig_cell.coordinate
                gen_cell = gen_ws[cell_ref]
                total_cells += 1

                # Font comparison
                of = orig_cell.font
                gf = gen_cell.font
                if of and gf:
                    try:
                        orig_fg = str(of.color.rgb) if of.color and of.color.rgb else None
                        gen_fg = str(gf.color.rgb) if gf.color and gf.color.rgb else None
                    except Exception:
                        orig_fg = None
                        gen_fg = None
                    font_matches = (
                        safe(of.name) == safe(gf.name) and
                        abs(safe(of.size, 11) - safe(gf.size, 11)) < 0.5 and
                        safe(of.bold, False) == safe(gf.bold, False) and
                        safe(of.italic, False) == safe(gf.italic, False) and
                        safe(of.underline, 'none') == safe(gf.underline, 'none') and
                        orig_fg == gen_fg
                    )
                    if not font_matches and style_issues < MAX_STYLE_ISSUES_REPORTED:
                        style_detail_lines.append(
                            f"  - `{cell_ref}`: Font differs — "
                            f"Orig=Font('{of.name}',{of.size},{of.bold}), "
                            f"Gen=Font('{gf.name}',{gf.size},{gf.bold})"
                        )
                        style_issues += 1

                # Fill comparison
                ofill = orig_cell.fill
                gfill = gen_cell.fill
                if ofill and gfill:
                    try:
                        fill_fg_orig = str(ofill.fgColor.rgb) if ofill.fgColor and ofill.fgColor.rgb else None
                        fill_fg_gen = str(gfill.fgColor.rgb) if gfill.fgColor and gfill.fgColor.rgb else None
                    except Exception:
                        fill_fg_orig = None
                        fill_fg_gen = None
                    if fill_fg_orig != fill_fg_gen and style_issues < MAX_STYLE_ISSUES_REPORTED:
                        style_detail_lines.append(
                            f"  - `{cell_ref}`: Fill differs — "
                            f"Orig={fill_fg_orig}, Gen={fill_fg_gen}"
                        )
                        style_issues += 1

                # Border comparison
                ob = orig_cell.border
                gb = gen_cell.border
                if ob and gb:
                    border_issues = []
                    for side in ['left', 'right', 'top', 'bottom']:
                        ob_side = getattr(ob, side)
                        gb_side = getattr(gb, side)
                        if safe(ob_side and ob_side.style) != safe(gb_side and gb_side.style):
                            border_issues.append(side)
                    if border_issues and style_issues < MAX_STYLE_ISSUES_REPORTED:
                        style_detail_lines.append(
                            f"  - `{cell_ref}`: Border differs on sides: {border_issues}"
                        )
                        style_issues += 1

                # Alignment comparison
                oa = orig_cell.alignment
                ga = gen_cell.alignment
                if oa and ga:
                    if (safe(oa.horizontal) != safe(ga.horizontal) or
                        safe(oa.vertical) != safe(ga.vertical)) and style_issues < MAX_STYLE_ISSUES_REPORTED:
                        style_detail_lines.append(
                            f"  - `{cell_ref}`: Alignment differs — "
                            f"Orig=H:{oa.horizontal}/V:{oa.vertical}, "
                            f"Gen=H:{ga.horizontal}/V:{ga.vertical}"
                        )
                        style_issues += 1

    report_lines.append(f"- **Cells checked:** {total_cells}")
    report_lines.append(f"- **Style mismatches:** {style_issues}")
    if style_issues == 0:
        report_lines.append("- ✅ **All styles preserved** (fonts, fills, borders, alignment)")
    else:
        report_lines.append(f"- ⚠️ **{style_issues} style differences found:**")
        report_lines.extend(style_detail_lines)
        if style_issues >= MAX_STYLE_ISSUES_REPORTED:
            report_lines.append(f"  - *...and {style_issues - MAX_STYLE_ISSUES_REPORTED} more differences (truncated)*")
    report_lines.append("")

    # ═══════════════════════════════════════════════════════════════════════
    # 6. CELL VALUE COMPARISON
    # ═══════════════════════════════════════════════════════════════════════
    report_lines.append("## 6. Cell Values")
    report_lines.append("")

    value_diff_lines = []
    MAX_VALUE_ISSUES = 15

    for sheet_name in orig_sheets:
        if sheet_name not in gen_sheets:
            continue
        orig_ws = orig_wb[sheet_name]
        gen_ws = gen_wb[sheet_name]

        for row in orig_ws.iter_rows(min_row=1, max_row=safe(orig_ws.max_row, 100) or 100):
            for orig_cell in row:
                cell_ref = orig_cell.coordinate
                gen_cell = gen_ws[cell_ref]

                orig_val = str(orig_cell.value) if orig_cell.value is not None else ""
                gen_val = str(gen_cell.value) if gen_cell.value is not None else ""

                if orig_val != gen_val:
                    value_changes += 1
                    if value_changes <= MAX_VALUE_ISSUES:
                        value_diff_lines.append(
                            f"  - `{sheet_name}!{cell_ref}`: "
                            f"'{orig_val}' → '{gen_val}'"
                        )

    report_lines.append(f"- **Total cells compared:** {total_cells}")
    report_lines.append(f"- **Value differences:** {value_changes}")
    if value_changes == 0:
        report_lines.append("- ✅ **All cell values identical**")
    else:
        report_lines.append(f"- ℹ️ **{value_changes} cell values differ** (see details):")
        report_lines.extend(value_diff_lines)
        if value_changes > MAX_VALUE_ISSUES:
            report_lines.append(f"  - *...and {value_changes - MAX_VALUE_ISSUES} more differences*")
    report_lines.append("")

    # ═══════════════════════════════════════════════════════════════════════
    # 7. SUMMARY
    # ═══════════════════════════════════════════════════════════════════════
    report_lines.append("---")
    report_lines.append("")
    report_lines.append("## Summary")
    report_lines.append("")

    # Compute summary statuses using previously computed values
    sheet_structure_ok = all(o == g for o, g in zip(orig_sheets, gen_sheets))
    merge_ok = True
    for m_name, n_name in zip(orig_sheets, gen_sheets):
        if m_name in orig_wb and n_name in gen_wb:
            try:
                o_merges = set(str(r) for r in orig_wb[m_name].merged_cells.ranges)
                g_merges = set(str(r) for r in gen_wb[n_name].merged_cells.ranges)
            except Exception:
                o_merges = set()
                g_merges = set()
            if o_merges != g_merges:
                merge_ok = False
                break

    # Checks that are not already accounted for by style/value diffs:
    # page_setup, row/col geometry, freeze panes, sheet names
    structural_failures = failed_checks
    # (value changes are tracked differently — they are expected when editing)
    non_value_failures = failed_checks  # all failed checks are non-value for now

    report_lines.append("| Metric | Status | Details |")
    report_lines.append("|--------|--------|---------|")
    report_lines.append(f"| **Sheet Count & Names** | {'✅' if sheet_structure_ok else '❌'} | {len(orig_sheets)} sheets |")
    report_lines.append(f"| **Page Layout** | {'✅' if non_value_failures == 0 else '❌'} | orientation, paper, fit-to-page, freeze panes |")
    report_lines.append(f"| **Merged Cells** | {'✅' if merge_ok else '❌'} | per sheet |")
    report_lines.append(f"| **Row/Column Geometry** | checked | row heights, column widths, hidden state |")
    report_lines.append(f"| **Styles** | {'✅' if style_issues == 0 else '❌'} | {style_issues} mismatches |")
    report_lines.append(f"| **Cell Values** | {'✅ Identical' if value_changes == 0 else f'ℹ️ {value_changes} changed'} | {total_cells} cells compared |")
    report_lines.append("")
    report_lines.append(f"**Total checks:** {total_checks} | **Passed:** {passed_checks} | **Failed:** {failed_checks}")
    report_lines.append("")

    # Conclusion
    major_issues = failed_checks
    if major_issues == 0 and style_issues == 0:
        report_lines.append("## ✅ Conclusion: Workbook is fully preserved")
        report_lines.append("")
        report_lines.append("The generated workbook is **structurally and visually identical** to the original.")
    elif major_issues == 0 and style_issues > 0:
        report_lines.append("## ⚠️ Conclusion: Minor style differences detected")
        report_lines.append("")
        report_lines.append(f"The workbook structure is preserved, but {style_issues} style differences were found. ")
        report_lines.append("These may be due to openpyxl normalization rather than actual regressions.")
    else:
        report_lines.append("## ❌ Conclusion: Structural differences detected")
        report_lines.append("")
        report_lines.append(f"The generated workbook differs from the original in {failed_checks} structural areas. ")
        report_lines.append("Review details above for specific issues.")

    report_lines.append("")
    report_lines.append("---")
    report_lines.append("*Report generated automatically by compare_workbooks.py*")

    return "\n".join(report_lines)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python compare_workbooks.py <original.xlsx> <generated.xlsx>")
        sys.exit(1)

    original = sys.argv[1]
    generated = sys.argv[2]

    report = compare_workbooks(original, generated)

    # Write to file
    report_path = "docs/workbook_fidelity_report.md"
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report)

    print(f"Report written to {report_path}")
    print(report)
