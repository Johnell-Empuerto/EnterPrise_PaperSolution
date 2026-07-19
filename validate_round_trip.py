#!/usr/bin/env python3
"""
Phase 4.6 — End-to-End Workbook Fidelity Validation Script.

Validates that the edited workbook preserves ALL original formatting,
structure, and layout — only cell values should differ.

Usage:
    python validate_round_trip.py <original.xlsx> <edited.xlsx> [--verbose]

Example:
    python validate_round_trip.py "FormTest - Copy.xlsx" "FormTest - Copy_output (11).xlsx"
"""

import sys
import hashlib
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict


class WorkbookFidelityValidator:
    """Validates workbook fidelity between original and edited files."""

    def __init__(self, original_path: str, edited_path: str, verbose: bool = False):
        self.orig_path = Path(original_path)
        self.edit_path = Path(edited_path)
        self.verbose = verbose
        self.results = {
            "workbook": {"pass": True, "checks": {}},
            "sheets": {},
            "summary": {
                "total_differences": 0,
                "value_changes": 0,
                "structural_changes": 0,
            },
        }

    def log(self, msg: str, level: str = "info"):
        if self.verbose or level in ("error", "warn", "result"):
            prefix = {"info": "  ", "warn": "  ⚠ ", "error": "  ❌ ", "result": "  ✅ "}
            print(f"{prefix.get(level, '  ')}{msg}")

    def sha256(self, path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def run(self) -> dict:
        print(f"\n{'='*60}")
        print(f"  Workbook Fidelity Validation")
        print(f"{'='*60}")
        print(f"  Original: {self.orig_path.name}")
        print(f"  Edited:   {self.edit_path.name}")
        print(f"{'='*60}\n")

        # 1. File-level checks
        self._check_file_hashes()

        # Open both workbooks
        wb_orig = load_workbook(self.orig_path, data_only=False)
        wb_edit = load_workbook(self.edit_path, data_only=False)

        # 2. Workbook-level checks
        self._check_workbook_level(wb_orig, wb_edit)

        # 3. Sheet-level checks
        for sheet_name in wb_orig.sheetnames:
            if sheet_name not in wb_edit.sheetnames:
                self._record_failure(
                    "sheets", sheet_name, "sheet_missing",
                    f"Sheet '{sheet_name}' missing in edited workbook"
                )
                continue

            ws_orig = wb_orig[sheet_name]
            ws_edit = wb_edit[sheet_name]
            self._check_sheet(ws_orig, ws_edit, sheet_name)

        # 4. Summary
        self._print_summary()

        return self.results

    def _check_file_hashes(self):
        orig_hash = self.sha256(self.orig_path)
        edit_hash = self.sha256(self.edit_path)
        self.log(f"Original SHA256: {orig_hash}")
        self.log(f"Edited SHA256:   {edit_hash}")
        if orig_hash == edit_hash:
            self.log("WARNING: Files are identical — no values were edited!", "warn")
        else:
            self.log("Hashes differ — values were changed (expected)", "result")

    def _check_workbook_level(self, wb_orig, wb_edit):
        checks = self.results["workbook"]["checks"]

        # Sheet count
        orig_count = len(wb_orig.sheetnames)
        edit_count = len(wb_edit.sheetnames)
        if orig_count != edit_count:
            self._record_failure(
                "workbook", "global", "sheet_count",
                f"Sheet count: {orig_count} vs {edit_count}"
            )
        else:
            checks["sheet_count_match"] = True
            self.log(f"Sheet count: {orig_count} ✅", "result")

    def _check_sheet(self, ws_orig, ws_edit, sheet_name: str):
        sheet_result = {
            "pass": True,
            "structural_changes": 0,
            "value_changes": 0,
            "checks": {},
            "changed_cells": [],
        }

        # Dimensions
        if ws_orig.dimensions != ws_edit.dimensions:
            sheet_result["structural_changes"] += 1
            self._record_failure(
                "sheets", sheet_name, "dimensions",
                f"Dimensions: {ws_orig.dimensions} vs {ws_edit.dimensions}"
            )

        # Row count
        orig_rows = len(list(ws_orig.iter_rows()))
        edit_rows = len(list(ws_edit.iter_rows()))
        if orig_rows != edit_rows:
            sheet_result["structural_changes"] += 1
            self._record_failure(
                "sheets", sheet_name, "row_count",
                f"Row count: {orig_rows} vs {edit_rows}"
            )
        else:
            sheet_result["checks"]["row_count_match"] = True

        # Column count
        orig_cols = ws_orig.max_column
        edit_cols = ws_edit.max_column
        if orig_cols != edit_cols:
            sheet_result["structural_changes"] += 1
            self._record_failure(
                "sheets", sheet_name, "column_count",
                f"Column count: {orig_cols} vs {edit_cols}"
            )
        else:
            sheet_result["checks"]["column_count_match"] = True

        # Merged cells
        orig_merges = set(ws_orig.merged_cells.ranges)
        edit_merges = set(ws_edit.merged_cells.ranges)
        if orig_merges != edit_merges:
            sheet_result["structural_changes"] += 1
            added = edit_merges - orig_merges
            removed = orig_merges - edit_merges
            self._record_failure(
                "sheets", sheet_name, "merged_cells",
                f"Merged cells differ: +{len(added)} -{len(removed)}"
            )
            if added:
                self.log(f"  Added merges: {added}", "warn")
            if removed:
                self.log(f"  Removed merges: {removed}", "error")
        else:
            sheet_result["checks"]["merged_cells_match"] = True

        # Print area
        if ws_orig.print_area != ws_edit.print_area:
            sheet_result["structural_changes"] += 1
            self._record_failure(
                "sheets", sheet_name, "print_area",
                f"Print area: {ws_orig.print_area} vs {ws_edit.print_area}"
            )
        else:
            sheet_result["checks"]["print_area_match"] = True

        # Page setup (orientation, paper size)
        orig_ps = ws_orig.page_setup
        edit_ps = ws_edit.page_setup
        ps_match = (
            orig_ps.orientation == edit_ps.orientation
            and orig_ps.paperSize == edit_ps.paperSize
            and orig_ps.fitToWidth == edit_ps.fitToWidth
            and orig_ps.fitToHeight == edit_ps.fitToHeight
        )
        if not ps_match:
            sheet_result["structural_changes"] += 1
            self._record_failure(
                "sheets", sheet_name, "page_setup",
                f"Page setup differs"
            )
        else:
            sheet_result["checks"]["page_setup_match"] = True

        # Page margins
        orig_mg = ws_orig.page_margins
        edit_mg = ws_edit.page_margins
        if orig_mg and edit_mg:
            mg_match = (
                orig_mg.left == edit_mg.left
                and orig_mg.right == edit_mg.right
                and orig_mg.top == edit_mg.top
                and orig_mg.bottom == edit_mg.bottom
                and orig_mg.header == edit_mg.header
                and orig_mg.footer == edit_mg.footer
            )
            if not mg_match:
                sheet_result["structural_changes"] += 1
                self._record_failure(
                    "sheets", sheet_name, "page_margins",
                    f"Page margins differ"
                )
            else:
                sheet_result["checks"]["page_margins_match"] = True

        # Freeze panes
        if ws_orig.freeze_panes != ws_edit.freeze_panes:
            sheet_result["structural_changes"] += 1
            self._record_failure(
                "sheets", sheet_name, "freeze_panes",
                f"Freeze panes: {ws_orig.freeze_panes} vs {ws_edit.freeze_panes}"
            )
        else:
            sheet_result["checks"]["freeze_panes_match"] = True

        # Row dimensions (heights + hidden)
        orig_rows_info = {
            r: {"height": ws_orig.row_dimensions[r].height,
                "hidden": ws_orig.row_dimensions[r].hidden}
            for r in ws_orig.row_dimensions
        }
        edit_rows_info = {
            r: {"height": ws_edit.row_dimensions[r].height,
                "hidden": ws_edit.row_dimensions[r].hidden}
            for r in ws_edit.row_dimensions
        }

        for r, info in orig_rows_info.items():
            if r not in edit_rows_info:
                sheet_result["structural_changes"] += 1
                self._record_failure(
                    "sheets", sheet_name, "row_missing",
                    f"Row {r} missing in edited workbook"
                )
            else:
                e_info = edit_rows_info[r]
                if info["height"] != e_info["height"]:
                    sheet_result["structural_changes"] += 1
                    self._record_failure(
                        "sheets", sheet_name, "row_height",
                        f"Row {r} height: {info['height']} vs {e_info['height']}"
                    )
                if info["hidden"] != e_info["hidden"]:
                    sheet_result["structural_changes"] += 1
                    self._record_failure(
                        "sheets", sheet_name, "row_hidden",
                        f"Row {r} hidden: {info['hidden']} vs {e_info['hidden']}"
                    )

        # Column dimensions (widths + hidden)
        orig_cols_info = {
            c: {"width": ws_orig.column_dimensions[c].width,
                "hidden": ws_orig.column_dimensions[c].hidden}
            for c in ws_orig.column_dimensions
        }
        edit_cols_info = {
            c: {"width": ws_edit.column_dimensions[c].width,
                "hidden": ws_edit.column_dimensions[c].hidden}
            for c in ws_edit.column_dimensions
        }

        for c, info in orig_cols_info.items():
            if c not in edit_cols_info:
                sheet_result["structural_changes"] += 1
                self._record_failure(
                    "sheets", sheet_name, "col_missing",
                    f"Col {c} missing in edited workbook"
                )
            else:
                e_info = edit_cols_info[c]
                if info["width"] != e_info["width"]:
                    sheet_result["structural_changes"] += 1
                    self._record_failure(
                        "sheets", sheet_name, "col_width",
                        f"Col {c} width: {info['width']} vs {e_info['width']}"
                    )
                if info["hidden"] != e_info["hidden"]:
                    sheet_result["structural_changes"] += 1
                    self._record_failure(
                        "sheets", sheet_name, "col_hidden",
                        f"Col {c} hidden: {info['hidden']} vs {e_info['hidden']}"
                    )

        # Cell-by-cell comparison
        all_cells = set()
        for row in ws_orig.iter_rows():
            for cell in row:
                all_cells.add(cell.coordinate)
        for row in ws_edit.iter_rows():
            for cell in row:
                all_cells.add(cell.coordinate)

        changed_cells = []
        for coord in sorted(all_cells, key=lambda x: (int(''.join(filter(str.isdigit, x)) or 0), ''.join(filter(str.isalpha, x)))):
            c_orig = ws_orig[coord]
            c_edit = ws_edit[coord]

            # Compare value
            v_orig = c_orig.value
            v_edit = c_edit.value
            if v_orig != v_edit:
                changed_cells.append(f"{coord}: '{v_orig}' → '{v_edit}'")
                sheet_result["value_changes"] += 1
                continue

            # Compare style (only if value is same — if value differs, style is expected to remain)
            if c_orig.font != c_edit.font:
                sheet_result["structural_changes"] += 1
                self._record_failure(
                    "sheets", sheet_name, "font",
                    f"Cell {coord}: font differs (value unchanged)"
                )
            if c_orig.fill != c_edit.fill:
                sheet_result["structural_changes"] += 1
                self._record_failure(
                    "sheets", sheet_name, "fill",
                    f"Cell {coord}: fill differs (value unchanged)"
                )
            if c_orig.border != c_edit.border:
                sheet_result["structural_changes"] += 1
                self._record_failure(
                    "sheets", sheet_name, "border",
                    f"Cell {coord}: border differs (value unchanged)"
                )
            if c_orig.alignment != c_edit.alignment:
                sheet_result["structural_changes"] += 1
                self._record_failure(
                    "sheets", sheet_name, "alignment",
                    f"Cell {coord}: alignment differs (value unchanged)"
                )
            if c_orig.number_format != c_edit.number_format:
                sheet_result["structural_changes"] += 1
                self._record_failure(
                    "sheets", sheet_name, "number_format",
                    f"Cell {coord}: number_format differs (value unchanged)"
                )

        if changed_cells and self.verbose:
            self.log(f"\n  Changed cells in '{sheet_name}':")
            for cc in changed_cells:
                self.log(f"    {cc}")

        sheet_result["changed_cells"] = changed_cells
        sheet_result["pass"] = sheet_result["structural_changes"] == 0

        # Print sheet-level result
        if sheet_result["structural_changes"] == 0:
            self.log(
                f"Sheet '{sheet_name}': {sheet_result['value_changes']} value changes, "
                f"0 structural changes ✅",
                "result"
            )
        else:
            self.log(
                f"Sheet '{sheet_name}': {sheet_result['value_changes']} value changes, "
                f"{sheet_result['structural_changes']} structural changes ❌",
                "error"
            )

        self.results["sheets"][sheet_name] = sheet_result
        self.results["summary"]["value_changes"] += sheet_result["value_changes"]
        self.results["summary"]["structural_changes"] += sheet_result["structural_changes"]

    def _record_failure(self, category: str, name: str, check: str, message: str):
        self.log(f"FAILURE [{category}/{name}]: {message}", "error")
        self.results["summary"]["total_differences"] += 1
        if category == "workbook":
            self.results["workbook"]["pass"] = False
            self.results["workbook"]["checks"][check] = message
        elif category == "sheets":
            if name not in self.results["sheets"]:
                self.results["sheets"][name] = {
                    "pass": False, "structural_changes": 0, "value_changes": 0,
                    "checks": {}, "changed_cells": []
                }
            self.results["sheets"][name]["pass"] = False

    def _print_summary(self):
        s = self.results["summary"]
        passed = s["structural_changes"] == 0

        print(f"\n{'='*60}")
        print(f"  VALIDATION SUMMARY")
        print(f"{'='*60}")
        print(f"  Editable Value Changes: {s['value_changes']}")
        print(f"  Structural Changes:     {s['structural_changes']}")
        print(f"  Total Differences:      {s['total_differences']}")
        print(f"")
        if passed:
            print(f"  🟢 RESULT: PASS")
        else:
            print(f"  🔴 RESULT: FAIL")
        print(f"{'='*60}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Validate workbook fidelity between original and edited XLSX files."
    )
    parser.add_argument("original", help="Path to the original workbook")
    parser.add_argument("edited", help="Path to the edited workbook")
    parser.add_argument("--verbose", "-v", action="store_true", help="Show detailed output")
    args = parser.parse_args()

    if not Path(args.original).exists():
        print(f"❌ Original file not found: {args.original}")
        sys.exit(1)
    if not Path(args.edited).exists():
        print(f"❌ Edited file not found: {args.edited}")
        sys.exit(1)

    validator = WorkbookFidelityValidator(args.original, args.edited, verbose=args.verbose)
    results = validator.run()

    if results["summary"]["structural_changes"] > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
