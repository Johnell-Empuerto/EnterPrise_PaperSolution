#!/usr/bin/env python3
"""
Phase 5.0 — Regression Workbook Generator.

Generates a comprehensive set of test workbooks covering every workbook
feature category for production certification testing.

Usage:
    python generate_workbooks.py [output_dir]

Default output directory: RegressionWorkbooks/
"""

import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule


OUTPUT_DIR = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("RegressionWorkbooks")


def make_workbook(filename: str, description: str) -> Workbook:
    """Create a workbook with filename and return it."""
    print(f"  Generating: {filename} ({description})")
    return Workbook()


def save_workbook(wb: Workbook, filename: str):
    """Save workbook to output directory."""
    output_path = OUTPUT_DIR / filename
    wb.save(str(output_path))
    print(f"    -> {output_path} ({output_path.stat().st_size} bytes)")


def generate_merged_cells():
    """Workbook with merged cells."""
    wb = make_workbook("01_merged_cells.xlsx", "merged cells in various patterns")
    ws = wb.active
    ws.title = "Merged"

    # Simple horizontal merge
    ws.merge_cells("A1:C1")
    ws["A1"] = "Horizontal Merge"
    ws["A1"].font = Font(bold=True, size=14)

    # Vertical merge
    ws.merge_cells("A3:A6")
    ws["A3"] = "Vertical Merge"
    ws["A3"].alignment = Alignment(vertical="center")

    # Block merge
    ws.merge_cells("C3:E6")
    ws["C3"] = "Block Merge"
    ws["C3"].alignment = Alignment(horizontal="center", vertical="center")

    # Multiple merge ranges
    ws.merge_cells("A8:C10")
    ws.merge_cells("D8:F10")
    ws.merge_cells("A12:F12")

    ws["A8"] = "Merge 1"
    ws["D8"] = "Merge 2"
    ws["A12"] = "Full Width Merge"

    # Overlapping merge areas (handled carefully)
    ws.merge_cells("A14:C16")
    ws.merge_cells("B15:D17")  # Different range

    save_workbook(wb, "01_merged_cells.xlsx")


def generate_hidden_rows_cols():
    """Workbook with hidden rows and columns."""
    wb = make_workbook("02_hidden_rows_cols.xlsx", "hidden rows and columns")
    ws = wb.active
    ws.title = "HiddenRC"

    # Fill data
    for row in range(1, 20):
        for col in range(1, 10):
            ws.cell(row=row, column=col, value=f"R{row}C{col}")

    # Hide specific rows
    ws.row_dimensions[3].hidden = True
    ws.row_dimensions[5].hidden = True
    ws.row_dimensions[8].hidden = True

    # Hide specific columns
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["D"].hidden = True
    ws.column_dimensions["G"].hidden = True

    # Mixed visible/hidden
    ws.row_dimensions[12].hidden = True
    ws.column_dimensions["F"].hidden = True

    save_workbook(wb, "02_hidden_rows_cols.xlsx")


def generate_formulas():
    """Workbook with formulas."""
    wb = make_workbook("03_formulas.xlsx", "various formula types")
    ws = wb.active
    ws.title = "Formulas"

    # Input values
    for row in range(1, 11):
        ws.cell(row=row, column=1, value=row * 10)
        ws.cell(row=row, column=2, value=row * 5)

    # SUM
    ws["D1"] = "=SUM(A1:A10)"
    ws["D2"] = "=SUM(B1:B10)"

    # AVERAGE
    ws["D4"] = "=AVERAGE(A1:A10)"
    ws["D5"] = "=AVERAGE(B1:B10)"

    # MIN / MAX
    ws["D7"] = "=MIN(A1:A10)"
    ws["D8"] = "=MAX(A1:A10)"

    # COUNT
    ws["D10"] = "=COUNT(A1:A10)"

    # IF statement
    ws["F1"] = "=IF(A1>50,\"High\",\"Low\")"
    ws["F2"] = "=IF(B1>30,\"High\",\"Low\")"

    # VLOOKUP
    ws["F4"] = '=VLOOKUP(50,A1:B10,2,FALSE)'

    # Concatenation
    ws["F6"] = '=CONCATENATE(A1," - ",B1)'

    # Date formulas
    ws["F8"] = "=TODAY()"
    ws["F9"] = "=NOW()"

    # Nested formulas
    ws["F11"] = "=IF(SUM(A1:A10)>100,\"Over 100\",\"Under 100\")"

    save_workbook(wb, "03_formulas.xlsx")


def generate_comments():
    """Workbook with cell comments."""
    wb = make_workbook("04_comments.xlsx", "cell comments on various cells")
    ws = wb.active
    ws.title = "Comments"

    for row in range(1, 15):
        ws.cell(row=row, column=1, value=f"Cell A{row}")

    ws["A1"].comment = "This is the first cell"
    ws["A3"].comment = "Important: review this value"
    ws["A5"].comment = "This field requires manager approval"
    ws["A7"].comment = "Auto-calculated, do not edit"
    ws["A9"].comment = "See Sheet2 for reference data"
    ws["A11"].comment = "Updated: 2026-01-15 by John"
    ws["A13"].comment = "Multi-line\ncomment\nexample"

    ws["C1"].comment = "Column header"
    ws["C3"].comment = "Input field"

    save_workbook(wb, "04_comments.xlsx")


def generate_conditional_formatting():
    """Workbook with conditional formatting."""
    wb = make_workbook("05_conditional_formatting.xlsx", "conditional formatting rules")
    ws = wb.active
    ws.title = "CondFormat"

    for row in range(1, 31):
        import random
        ws.cell(row=row, column=1, value=random.randint(1, 100))
        ws.cell(row=row, column=2, value=random.randint(1, 100))
        ws.cell(row=row, column=3, value=random.choice(["Pass", "Fail", "Review"]))

    # Highlight cells > 75
    ws.conditional_formatting.add(
        "A1:A30",
        CellIsRule(operator="greaterThan", formula=["75"],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    # Highlight cells < 25
    ws.conditional_formatting.add(
        "B1:B30",
        CellIsRule(operator="lessThan", formula=["25"],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )

    # Highlight "Fail" cells
    ws.conditional_formatting.add(
        "C1:C30",
        CellIsRule(operator="equal", formula=['"Fail"'],
                   font=Font(bold=True, color="FF0000"))
    )

    save_workbook(wb, "05_conditional_formatting.xlsx")


def generate_styles():
    """Workbook with comprehensive styling."""
    wb = make_workbook("06_styles.xlsx", "comprehensive styles: fonts, fills, borders, alignment")
    ws = wb.active
    ws.title = "Styles"

    # Font variations
    ws["A1"] = "Bold"
    ws["A1"].font = Font(bold=True, size=11)
    ws["B1"] = "Italic"
    ws["B1"].font = Font(italic=True, size=11)
    ws["C1"] = "Underline"
    ws["C1"].font = Font(underline="single", size=11)
    ws["D1"] = "Colored"
    ws["D1"].font = Font(color="FF0000", size=11)
    ws["E1"] = "Large"
    ws["E1"].font = Font(size=18)
    ws["F1"] = "Arial"
    ws["F1"].font = Font(name="Arial", size=11)

    # Fill variations
    colors = ["FFFF00", "00FF00", "0000FF", "FF00FF", "00FFFF"]
    for i, color in enumerate(colors):
        cell = ws.cell(row=3, column=i + 1, value=f"Fill {i + 1}")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Border variations
    thin = Side(style="thin")
    thick = Side(style="thick")
    double = Side(style="double")
    dashed = Side(style="dashed")
    dotted = Side(style="dotted")

    ws["A5"].value = "Thin"
    ws["A5"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["B5"].value = "Thick"
    ws["B5"].border = Border(left=thick, right=thick, top=thick, bottom=thick)
    ws["C5"].value = "Double"
    ws["C5"].border = Border(left=double, right=double, top=double, bottom=double)
    ws["D5"].value = "Dashed"
    ws["D5"].border = Border(left=dashed, right=dashed, top=dashed, bottom=dashed)
    ws["E5"].value = "Mixed"
    ws["E5"].border = Border(left=thin, right=thick, top=double, bottom=dashed)

    # Alignment variations
    alignments = [
        Alignment(horizontal="left", vertical="top"),
        Alignment(horizontal="center", vertical="center"),
        Alignment(horizontal="right", vertical="bottom"),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
        Alignment(horizontal="center", vertical="center", text_rotation=45),
        Alignment(horizontal="center", vertical="center", text_rotation=90),
    ]
    for i, align in enumerate(alignments):
        cell = ws.cell(row=7, column=i + 1, value=f"Align {i + 1}")
        cell.alignment = align
        if i == 3:
            cell.value = "Wrapped text example with longer content"

    # Number formats
    formats = [
        '#,##0',  # Number with commas
        '#,##0.00',  # Decimal
        '$#,##0.00',  # Currency
        '0%',  # Percentage
        '0.00%',  # Percentage decimal
        'mm/dd/yyyy',  # Date
        'dddd, mmmm d, yyyy',  # Long date
        'h:mm:ss AM/PM',  # Time
    ]
    for i, nf in enumerate(formats):
        cell = ws.cell(row=9, column=i + 1, value=1234.56)
        cell.number_format = nf

    save_workbook(wb, "06_styles.xlsx")


def generate_print_settings():
    """Workbook with various print settings."""
    wb = make_workbook("07_print_settings.xlsx", "various page layouts and print areas")
    ws1 = wb.active
    ws1.title = "Portrait"

    for row in range(1, 51):
        for col in range(1, 10):
            ws1.cell(row=row, column=col, value=f"R{row}C{col}")

    ws1.page_setup.orientation = "portrait"
    ws1.page_setup.paperSize = 1  # Letter
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 1
    ws1.print_area = "A1:J50"
    ws1.page_margins.left = 0.75
    ws1.page_margins.right = 0.75
    ws1.page_margins.top = 1.0
    ws1.page_margins.bottom = 1.0
    ws1.page_margins.header = 0.5
    ws1.page_margins.footer = 0.5
    ws1.sheet_properties.pageSetUpPr.fitToPage = True

    ws2 = wb.create_sheet("Landscape")
    for row in range(1, 51):
        for col in range(1, 15):
            ws2.cell(row=row, column=col, value=f"R{row}C{col}")
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.paperSize = 3  # Tabloid
    ws2.page_setup.fitToWidth = 2
    ws2.page_setup.fitToHeight = 1
    ws2.print_area = "A1:N50"

    ws3 = wb.create_sheet("FitToPage")
    for row in range(1, 101):
        for col in range(1, 20):
            ws3.cell(row=row, column=col, value=f"R{row}C{col}")
    ws3.page_setup.orientation = "landscape"
    ws3.page_setup.fitToWidth = 1
    ws3.page_setup.fitToHeight = 0
    ws3.page_setup.zoom = 80
    ws3.sheet_properties.pageSetUpPr.fitToPage = True

    save_workbook(wb, "07_print_settings.xlsx")


def generate_multiple_sheets():
    """Workbook with many sheets."""
    wb = make_workbook("08_multiple_sheets.xlsx", "20+ worksheets")
    wb.active.title = "Sheet01"

    for row in range(1, 11):
        wb.active.cell(row=row, column=1, value=f"Sheet1-R{row}")

    for i in range(2, 21):
        ws = wb.create_sheet(f"Sheet{i:02d}")
        for row in range(1, 11):
            ws.cell(row=row, column=1, value=f"Sheet{i}-R{row}")

    save_workbook(wb, "08_multiple_sheets.xlsx")


def generate_named_ranges():
    """Workbook with many named ranges."""
    wb = make_workbook("09_named_ranges.xlsx", "defined names / named ranges")
    ws = wb.active
    ws.title = "Data"
    ws2 = wb.create_sheet("Lookup")

    for row in range(1, 21):
        ws.cell(row=row, column=1, value=row)
        ws.cell(row=row, column=2, value=f"Item {row}")
        ws.cell(row=row, column=3, value=row * 100)

    for row in range(1, 11):
        ws2.cell(row=row, column=1, value=row)
        ws2.cell(row=row, column=2, value=f"Lookup {row}")

    # Named ranges
    wb.defined_names["DataValues"] = "Data!$A$1:$A$20"
    wb.defined_names["ItemNames"] = "Data!$B$1:$B$20"
    wb.defined_names["Prices"] = "Data!$C$1:$C$20"
    wb.defined_names["LookupData"] = "Lookup!$A$1:$B$10"
    wb.defined_names["Header"] = "Data!$A$1:$C$1"
    wb.defined_names["TaxRate"] = "0.08"
    wb.defined_names["Discount"] = "0.10"

    save_workbook(wb, "09_named_ranges.xlsx")


def generate_freeze_panes():
    """Workbook with freeze panes."""
    wb = make_workbook("10_freeze_panes.xlsx", "freeze panes at various positions")
    ws1 = wb.active
    ws1.title = "FreezeRow1"
    for row in range(1, 31):
        for col in range(1, 10):
            ws1.cell(row=row, column=col, value=f"R{row}C{col}")
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("FreezeColA")
    for row in range(1, 31):
        for col in range(1, 10):
            ws2.cell(row=row, column=col, value=f"R{row}C{col}")
    ws2.freeze_panes = "B1"

    ws3 = wb.create_sheet("FreezeBoth")
    for row in range(1, 31):
        for col in range(1, 10):
            ws3.cell(row=row, column=col, value=f"R{row}C{col}")
    ws3.freeze_panes = "C4"

    save_workbook(wb, "10_freeze_panes.xlsx")


def generate_protected_sheets():
    """Workbook with sheet protection."""
    wb = make_workbook("11_protected_sheets.xlsx", "protected worksheets")
    ws = wb.active
    ws.title = "Protected"
    ws.protection.sheet = True
    ws.protection.password = "test123"
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.insertHyperlinks = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False

    for row in range(1, 11):
        for col in range(1, 6):
            ws.cell(row=row, column=col, value=f"R{row}C{col}")

    # Editable cells
    for row in range(1, 6):
        cell = ws.cell(row=row, column=3)
        cell.protection = Protection(locked=False)
        cell.value = ""

    # Unprotected sheet
    ws2 = wb.create_sheet("Unprotected")
    for row in range(1, 11):
        ws2.cell(row=row, column=1, value=f"R{row}")

    save_workbook(wb, "11_protected_sheets.xlsx")


def generate_data_validation():
    """Workbook with data validation rules."""
    wb = make_workbook("12_data_validation.xlsx", "data validation rules")
    ws = wb.active
    ws.title = "Validation"

    ws["A1"] = "List Validation"
    ws["B1"] = "Integer Range"
    ws["C1"] = "Decimal Range"
    ws["D1"] = "Date Range"
    ws["E1"] = "Text Length"

    # Dropdown list
    dv_list = DataValidation(
        type="list", formula1='"Option A,Option B,Option C,Option D"',
        allow_blank=True
    )
    dv_list.error = "Please select from the dropdown"
    dv_list.errorTitle = "Invalid Selection"
    ws.add_data_validation(dv_list)
    dv_list.add("A2:A20")

    # Integer range
    dv_int = DataValidation(type="whole", operator="between", formula1="1", formula2="100")
    dv_int.error = "Please enter a whole number between 1 and 100"
    ws.add_data_validation(dv_int)
    dv_int.add("B2:B20")

    # Decimal range
    dv_dec = DataValidation(type="decimal", operator="between", formula1="0.0", formula2="1.0")
    dv_dec.error = "Please enter a value between 0.0 and 1.0"
    ws.add_data_validation(dv_dec)
    dv_dec.add("C2:C20")

    # Date range
    dv_date = DataValidation(type="date", operator="greaterThan",
                             formula1="DATE(2024,1,1)")
    dv_date.error = "Please enter a date after 2024-01-01"
    ws.add_data_validation(dv_date)
    dv_date.add("D2:D20")

    # Text length
    dv_text = DataValidation(type="textLength", operator="lessThan", formula1="10")
    dv_text.error = "Please enter fewer than 10 characters"
    ws.add_data_validation(dv_text)
    dv_text.add("E2:E20")

    save_workbook(wb, "12_data_validation.xlsx")


def generate_unicode():
    """Workbook with Unicode/Japanese/Chinese/Korean text."""
    wb = make_workbook("13_unicode.xlsx", "Unicode Japanese/Chinese/Korean text")
    ws = wb.active
    ws.title = "Unicode"

    # Japanese
    ws["A1"] = "日本語"
    ws["A2"] = "こんにちは"
    ws["A3"] = "ありがとうございます"
    ws["A4"] = "おはようございます"

    # Chinese
    ws["C1"] = "中文"
    ws["C2"] = "你好"
    ws["C3"] = "谢谢"
    ws["C4"] = "早上好"

    # Korean
    ws["E1"] = "한국어"
    ws["E2"] = "안녕하세요"
    ws["E3"] = "감사합니다"
    ws["E4"] = "좋은 아침입니다"

    # Mixed
    ws["A6"] = "Hello こんにちは 你好 안녕하세요"
    ws["A7"] = "Mixed: English/日本語/中文/한국어"
    ws["A8"] = "123 ABC あいうえお 一二三 가나다"

    # Unicode in column headers
    for i, (label, text) in enumerate([
        ("日本語", "Japanese"),
        ("中文", "Chinese"),
        ("한국어", "Korean"),
    ]):
        cell = ws.cell(row=10, column=i * 2 + 1, value=f"{label} - {text}")
        cell.font = Font(bold=True, size=12)

    save_workbook(wb, "13_unicode.xlsx")


def generate_row_heights():
    """Workbook with varying row heights."""
    wb = make_workbook("14_row_heights.xlsx", "varying row heights and column widths")
    ws = wb.active
    ws.title = "Heights"

    for row in range(1, 16):
        ws.cell(row=row, column=1, value=f"Row {row}")
        ws.cell(row=row, column=2, value=f"Height varies")

    heights = [15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 30, 45, 20, 55, 100]
    for i, h in enumerate(heights):
        ws.row_dimensions[i + 1].height = h

    # Varying column widths
    widths = [8, 12, 15, 20, 25, 30, 35, 40, 50, 60]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    save_workbook(wb, "14_row_heights.xlsx")


def generate_all():
    """Generate all regression workbooks."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\n{'='*60}")
    print(f"  Generating Regression Workbooks")
    print(f"  Output: {OUTPUT_DIR.resolve()}")
    print(f"{'='*60}\n")

    generate_merged_cells()
    generate_hidden_rows_cols()
    generate_formulas()
    generate_comments()
    generate_conditional_formatting()
    generate_styles()
    generate_print_settings()
    generate_multiple_sheets()
    generate_named_ranges()
    generate_freeze_panes()
    generate_protected_sheets()
    generate_data_validation()
    generate_unicode()
    generate_row_heights()

    print(f"\n{'='*60}")
    print(f"  Generated {len(list(OUTPUT_DIR.glob('*.xlsx')))} workbooks")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    generate_all()
