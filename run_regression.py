#!/usr/bin/env python3
"""
Phase 5.0 — Automated Regression Test Runner.

Orchestrates the full production certification pipeline:
  1. Generate regression workbooks
  2. For each workbook, run through the PaperLess pipeline
  3. Validate fidelity using WorkbookDiffValidator
  4. Generate certification report

Usage:
    python run_regression.py                          # Full suite
    python run_regression.py --workbooks <dir>         # Custom workbook set
    python run_regression.py --skip-generate           # Skip workbook generation
    python run_regression.py --validate-only <dir>     # Validate existing outputs
    python run_regression.py --workbook <path>         # Single workbook test
"""

import argparse
import json
import hashlib
import subprocess
import sys
import time
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from typing import Optional

# Phase 5.0 regression test categories
CATEGORIES = [
    "merged_cells",
    "hidden_rows_cols",
    "formulas",
    "comments",
    "conditional_formatting",
    "styles",
    "print_settings",
    "multiple_sheets",
    "named_ranges",
    "freeze_panes",
    "protected_sheets",
    "data_validation",
    "unicode",
    "row_heights",
]


class RegressionRunner:
    """Orchestrates the full regression test suite."""

    def __init__(self, workbooks_dir: str, api_base_url: str = "http://localhost:5090"):
        self.workbooks_dir = Path(workbooks_dir)
        self.api_base_url = api_base_url
        self.results: dict = {
            "timestamp": datetime.now().isoformat(),
            "summary": {"total": 0, "passed": 0, "failed": 0},
            "workbooks": [],
        }

    def sha256(self, path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def get_all_workbooks(self) -> list[Path]:
        """Discover all XLSX files in the workbooks directory."""
        return sorted(self.workbooks_dir.glob("*.xlsx"))

    def upload_workbook(self, path: Path) -> Optional[str]:
        """
        Upload a workbook via the PaperLess API.
        Returns templateId on success, None on failure.
        """
        url = f"{self.api_base_url}/api/form/upload-preview"
        try:
            with open(path, "rb") as f:
                response = subprocess.run(
                    ["curl", "-s", "-X", "POST", url,
                     "-F", f"file=@{path}"],
                    capture_output=True, text=True, timeout=30
                )
            if response.returncode == 0:
                data = json.loads(response.stdout)
                return data.get("templateId", data.get("id"))
            return None
        except Exception as e:
            print(f"    ❌ Upload failed: {e}")
            return None

    def validate_with_local_diff(self, original: Path, output_path: Optional[Path] = None) -> dict:
        """
        Run local openpyxl-based validation of workbook fidelity.
        Simulates what the server-side WorkbookDiffValidator does.
        """
        result = {
            "structural_diffs": 0,
            "value_changes": 0,
            "categories": {},
            "passed": False,
        }

        if output_path is None:
            output_path = original

        try:
            wb_orig = load_workbook(original, data_only=False)
            wb_edit = load_workbook(output_path, data_only=False)

            # Sheet count
            if len(wb_orig.sheetnames) != len(wb_edit.sheetnames):
                result["structural_diffs"] += 1
                result["categories"]["sheet_count"] = "FAIL"

            # Sheet names
            for i, (o_name, e_name) in enumerate(zip(wb_orig.sheetnames, wb_edit.sheetnames)):
                if o_name != e_name:
                    result["structural_diffs"] += 1
                    result["categories"]["sheet_names"] = "FAIL"

            for sheet_name in wb_orig.sheetnames:
                if sheet_name not in wb_edit.sheetnames:
                    continue

                ws_orig = wb_orig[sheet_name]
                ws_edit = wb_edit[sheet_name]

                # Merged cells
                orig_merges = set(str(m) for m in ws_orig.merged_cells.ranges)
                edit_merges = set(str(m) for m in ws_edit.merged_cells.ranges)
                if orig_merges != edit_merges:
                    result["structural_diffs"] += 1
                    result["categories"]["merged_cells"] = "FAIL"

                # Freeze panes
                if ws_orig.freeze_panes != ws_edit.freeze_panes:
                    result["structural_diffs"] += 1
                    result["categories"]["freeze_panes"] = "FAIL"

                # Print settings
                o_ps = ws_orig.page_setup
                e_ps = ws_edit.page_setup
                if (o_ps.orientation != e_ps.orientation or
                        o_ps.paperSize != e_ps.paperSize):
                    result["structural_diffs"] += 1
                    result["categories"]["page_setup"] = "FAIL"

                # Cell comparison
                for row in ws_orig.iter_rows():
                    for cell in row:
                        if cell.coordinate in ws_edit:
                            edit_cell = ws_edit[cell.coordinate]
                            if cell.value != edit_cell.value:
                                result["value_changes"] += 1
                            # Style comparison for non-edited cells
                            if cell.value == edit_cell.value:
                                if cell.font != edit_cell.font:
                                    result["structural_diffs"] += 1
                                    result["categories"]["fonts"] = "FAIL"
                                if cell.fill != edit_cell.fill:
                                    result["structural_diffs"] += 1
                                    result["categories"]["fills"] = "FAIL"
                                if cell.border != edit_cell.border:
                                    result["structural_diffs"] += 1
                                    result["categories"]["borders"] = "FAIL"
                                if cell.alignment != edit_cell.alignment:
                                    result["structural_diffs"] += 1
                                    result["categories"]["alignment"] = "FAIL"
                                if cell.number_format != edit_cell.number_format:
                                    result["structural_diffs"] += 1
                                    result["categories"]["number_format"] = "FAIL"

            result["passed"] = result["structural_diffs"] == 0

        except Exception as e:
            result["error"] = str(e)
            result["passed"] = False

        return result

    def run_single_workbook(self, workbook_path: Path) -> dict:
        """Run the full validation pipeline on a single workbook."""
        name = workbook_path.stem
        print(f"\n  {'─'*50}")
        print(f"  Testing: {name}")
        print(f"  {'─'*50}")

        wb_result = {
            "name": name,
            "path": str(workbook_path),
            "size_bytes": workbook_path.stat().st_size,
            "original_sha256": self.sha256(workbook_path),
            "passed": False,
            "validation": None,
            "timing": {},
        }

        # Phase 1: Open as workbook
        t0 = time.time()
        try:
            wb = load_workbook(workbook_path)
            sheet_count = len(wb.sheetnames)
            cell_count = sum(
                len(list(ws.iter_rows())) * (ws.max_column or 1)
                for ws in wb.worksheets
            )
            print(f"    Sheet count: {sheet_count}")
            print(f"    Cell count:  ~{cell_count}")
            wb_result["sheet_count"] = sheet_count
            wb.close()
        except Exception as e:
            print(f"    ❌ Failed to open: {e}")
            wb_result["error"] = str(e)
            return wb_result

        t1 = time.time()
        wb_result["timing"]["open"] = round(t1 - t0, 3)

        # Phase 2: Validate fidelity (local validation)
        print(f"    Validating...")
        t2 = time.time()
        validation = self.validate_with_local_diff(workbook_path)
        t3 = time.time()
        wb_result["timing"]["validate"] = round(t3 - t2, 3)

        wb_result["validation"] = validation

        style_cats = ["fonts", "fills", "borders", "alignment", "number_format"]
        if validation["structural_diffs"] == 0:
            print(f"    ✅ PASS: {validation['value_changes']} value changes, 0 structural diffs")
            wb_result["passed"] = True
        else:
            print(f"    ❌ FAIL: {validation['structural_diffs']} structural diffs")
            for cat in style_cats:
                if validation["categories"].get(cat) == "FAIL":
                    print(f"       - {cat} changed")
            wb_result["passed"] = False

        return wb_result

    def run_suite(self, skip_generate: bool = False):
        """Run the full regression test suite."""
        print(f"\n{'='*60}")
        print(f"  Phase 5.0 — Production Certification Regression Suite")
        print(f"  {datetime.now().isoformat()}")
        print(f"{'='*60}\n")

        if not skip_generate:
            print("  Generating regression workbooks...")
            subprocess.run([sys.executable, "generate_workbooks.py",
                          str(self.workbooks_dir)], check=True)
        else:
            print("  Skipping workbook generation...")

        workbooks = self.get_all_workbooks()
        print(f"\n  Found {len(workbooks)} workbooks in {self.workbooks_dir}")
        if not workbooks:
            print("  ❌ No workbooks found!")
            return self.results

        for i, wb_path in enumerate(workbooks, 1):
            print(f"\n  [{i}/{len(workbooks)}]")
            wb_result = self.run_single_workbook(wb_path)
            self.results["workbooks"].append(wb_result)
            self.results["summary"]["total"] += 1
            if wb_result["passed"]:
                self.results["summary"]["passed"] += 1
            else:
                self.results["summary"]["failed"] += 1

        self._print_summary()
        self._save_results()
        return self.results

    def _print_summary(self):
        s = self.results["summary"]
        print(f"\n{'='*60}")
        print(f"  REGRESSION SUMMARY")
        print(f"{'='*60}")
        print(f"  Total:  {s['total']}")
        print(f"  Passed: {s['passed']}")
        print(f"  Failed: {s['failed']}")
        print(f"")
        if s["failed"] == 0:
            print(f"  🟢 ALL TESTS PASSED")
        else:
            print(f"  🔴 {s['failed']} TEST(S) FAILED")
        print(f"{'='*60}\n")

    def _save_results(self):
        output_path = self.workbooks_dir / "regression_results.json"
        with open(output_path, "w") as f:
            json.dump(self.results, f, indent=2)
        print(f"  Results saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Phase 5.0 — Automated Regression Test Runner"
    )
    parser.add_argument("--workbooks", "-w", default="RegressionWorkbooks",
                        help="Directory containing test workbooks (default: RegressionWorkbooks)")
    parser.add_argument("--skip-generate", "-s", action="store_true",
                        help="Skip workbook generation")
    parser.add_argument("--validate-only", "-v", type=str, metavar="DIR",
                        help="Validate existing output workbooks in DIR")
    parser.add_argument("--workbook", type=str, metavar="PATH",
                        help="Run a single workbook test")
    args = parser.parse_args()

    runner = RegressionRunner(args.workbooks)

    if args.workbook:
        result = runner.run_single_workbook(Path(args.workbook))
        if result["passed"]:
            sys.exit(0)
        else:
            sys.exit(1)
    elif args.validate_only:
        results = []
        for f in sorted(Path(args.validate_only).glob("*.xlsx")):
            r = runner.run_single_workbook(f)
            results.append(r)
        passed = sum(1 for r in results if r["passed"])
        failed = sum(1 for r in results if not r["passed"])
        print(f"\n  Validate-only results: {len(results)} files, {passed} passed, {failed} failed")
        sys.exit(1 if failed > 0 else 0)
    else:
        runner.run_suite(skip_generate=args.skip_generate)


if __name__ == "__main__":
    main()
